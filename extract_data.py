import polars as pl
import openpyxl
from pathlib import Path

# -------------------------------------------------------------------------------------------------------------#
#                                     Production values
#-------------------------------------------------------------------------------------------------------------#
excel_file = Path("RawData/Matriz_Prod_2023.xlsx").expanduser()

print(f"Cargando archivo Excel: {excel_file}")
workbook = openpyxl.load_workbook(excel_file, data_only=True)

sheet = workbook.active
print(f"Hoja de trabajo: {sheet.title}")

actividades = []
descripciones = []

ultima_columna = 124

for col in range(3, ultima_columna + 1):  # C es la columna 3
    codigo = sheet.cell(row=5, column=col).value
    descripcion = sheet.cell(row=6, column=col).value
    
    if codigo is not None:
        actividades.append(codigo)
        descripciones.append(descripcion)

valores_produccion = []

for col in range(3, ultima_columna + 1):
    valor = sheet.cell(row=403, column=col).value
    
    if valor is None:
        valor = 0
    
    valores_produccion.append(valor)

data = {
    "Código Actividad": actividades,
    "Descripción": descripciones,
    "Valor Producción": valores_produccion
}

df = pl.DataFrame(data)

# --------------------------------------------------------------#
#              Correlatives for the production values  
# --------------------------------------------------------------#

correlative_file = Path("ProcessedData/Correlativa 68 a 27.xlsx").expanduser()
correlative_df = pl.read_excel(correlative_file)


print("\nColumnas en datos de producción:", df.columns)
print("Columnas en datos de correlativa:", correlative_df.columns)


print("\nRealizando merge 1:1...")
merged_df = df.join(
    correlative_df,
    left_on="Código Actividad",
    right_on="Rama de Actividad",
    how="left"
)
print(f"\nResultado del merge: {len(merged_df)} filas")
print(merged_df.head(5))

print("\nCreando resumen por Act27...")
resume_act27 = merged_df.group_by("Act27").agg(
        pl.sum("Valor Producción").alias("Valor Producción Total")
    ).sort("Act27")

# -------------------------------------------------------------------------------------------------------------#
#                                     Empelados y Ocupados
#-------------------------------------------------------------------------------------------------------------#

excel_file = Path("Multiplicadores de Empleo Turismo - 2.1.xlsx").expanduser()

print(f"Loading Excel file: {excel_file}")
workbook = openpyxl.load_workbook(excel_file, data_only=True)

def extract_sheet_data(sheet_name):
    print(f"Extracting data from sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    
    data = []
    
    years = [str(year) for year in range(2015, 2025) if year != 2020] # Years 2015 to 2024, excluding 2020

    for row in range(13, 40):  # Rows 13 to 39
        row_data = {}
        
        activity_code = sheet.cell(row=row, column=1).value
        description = sheet.cell(row=row, column=2).value
        
        if activity_code is None:
            continue  # Skip empty rows
            
        row_data["Actividad Económica"] = activity_code
        row_data["Descripción"] = description
        
        for i, col in enumerate(range(3, len(years)+3)):  # Columns C to J
            if i < len(years):
                year = years[i]
                value = sheet.cell(row=row, column=col).value
                row_data[year] = value
        
        data.append(row_data)
    
    return data

sheet_names = workbook.sheetnames


# Production values are in thousands of pesos

empleos_data = extract_sheet_data("Empleos")
empleos_df = pl.DataFrame(empleos_data)
ocupados_data = extract_sheet_data("Ocupados")
ocupados_df = pl.DataFrame(ocupados_data)
production_df = resume_act27

print("Data extraction empleos.")
print(f"Empleos DataFrame shape: {empleos_df.head(5)}")
print(f"Ocupados DataFrame shape: {ocupados_df.head(5)}")
print(f"Production DataFrame shape: {resume_act27.head(5)}")
print(f"Total Production : {production_df['Valor Producción Total'].sum()}")


