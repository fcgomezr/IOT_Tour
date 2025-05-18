import polars as pl
import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import extract_data as ed
import numpy as np
import tensorflow as tf

# --------- loading the Ocup and TC ------------- # 
emp = ed.empleos_df.select(ed.empleos_df.columns[-1]).to_series() 
oci = ed.ocupados_df.select(ed.ocupados_df.columns[-1]).to_series() 
prod = ed.production_df.select(ed.production_df.columns[-1]).to_series()
emp = emp.to_numpy()
oci = oci.to_numpy()
prod = prod.to_numpy()


# ---------- load the IOT data ------------- #
dir = "ProcessedData/2019/"

ruta_leon = dir + 'Leon.csv'


df_leon = pl.read_csv(ruta_leon, separator=';', has_header=True, decimal_comma=True)
primera_columna = df_leon.columns[0]
matriz_leon = df_leon.drop(primera_columna).to_numpy()


g_tensor = tf.convert_to_tensor(prod, dtype=tf.float32)
leon_tensor = tf.convert_to_tensor(matriz_leon, dtype=tf.float32)
emp_tensor = tf.convert_to_tensor(emp, dtype=tf.float32)
oci_tensor = tf.convert_to_tensor(oci, dtype=tf.float32)


# --------- Model definition ------------- #
mod = tf.divide(emp_tensor, g_tensor)
mod2 = tf.divide(oci_tensor, g_tensor)
diag_oci = tf.linalg.diag(mod2)
diag_empl = tf.linalg.diag(mod) 

mele = tf.matmul(diag_empl, leon_tensor)   # Leon matrix EMPE
meloci = tf.matmul(diag_oci, leon_tensor)  # Leon matrix OCUP   
# --------- Model calculation ------------- #

nl = tf.linalg.inv(meloci)

emp_col = tf.expand_dims(oci_tensor, axis=-1)
F = tf.matmul( nl,emp_col)

pro = tf.matmul(meloci, F) - emp_col

## Copy the result to a new excel file

# Path to your Excel file
file_path = "Multiplicadores de Empleo Turismo - 2.1.xlsx"

# Assuming you already have the tensors F, meloci, and mele from your model
# Convert tensors to NumPy arrays for easier handling
F_np = F.numpy()
meloci_np = meloci.numpy()
mele_np = mele.numpy()

# Load the workbook
workbook = load_workbook(file_path)

# 1. Copy F to "Demanda Final" sheet, cells K13 to K39
sheet_df = workbook["Demanda Final"]
for i in range(F_np.shape[0]):
    # F_np is likely a column vector, so we take the first element of each row
    value = F_np[i, 0] if F_np.ndim > 1 else F_np[i]
    sheet_df.cell(row=i+13, column=11).value = value  # Column K is 11th column

# 2. Copy meloci to "Mult Ocup 2024" sheet, cells D13 to AD39
sheet_ocup = workbook["Mult Ocup 2024"]
for i in range(meloci_np.shape[0]):
    for j in range(meloci_np.shape[1]):
        # Only proceed if within our target range
        if i < 27 and j < 27:  # Assuming we have 27 rows and 27 columns to copy
            sheet_ocup.cell(row=i+13, column=j+4).value = meloci_np[i, j]  # Column D is 4th column

# 3. Copy mele to "Mult Empl 2024" sheet, cells D13 to AD39
sheet_empl = workbook["Mult Empl 2024"]
for i in range(mele_np.shape[0]):
    for j in range(mele_np.shape[1]):
        # Only proceed if within our target range
        if i < 27 and j < 27:  # Assuming we have 27 rows and 27 columns to copy
            sheet_empl.cell(row=i+13, column=j+4).value = mele_np[i, j]  # Column D is 4th column

# Save the workbook
workbook.save(file_path)
print(f"Data successfully copied to {file_path}")